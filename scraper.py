import json, re
from datetime import datetime
from playwright.sync_api import sync_playwright
import openpyxl
from io import BytesIO
import requests

URL = "https://www.banrep.gov.co/es/sen-boletines-diarios"

def scrape():
    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page()
        page.goto(URL, timeout=60000)
        page.wait_for_timeout(4000)
        links = page.eval_on_selector_all(
            'a[href*=".xlsx"], a[href*=".xls"]',
            'els => els.map(e => e.href)'
        )
        browser.close()

    if not links:
        print("No se encontraron links de Excel")
        return None

    url = links[0]
    print(f"Descargando: {url}")
    resp = requests.get(url, timeout=30)
    wb = openpyxl.load_workbook(BytesIO(resp.content), data_only=True)

    tes_data = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            label = str(row[0] or row[1] or '').strip()
            if 'TES TF' not in label.upper():
                continue
            tir = None
            for cell in row[1:]:
                if isinstance(cell, (int, float)) and 5 < cell < 25:
                    tir = round(float(cell), 4)
                    break
            if not tir:
                continue
            year_match = re.search(r'20(\d{2})', label)
            if not year_match:
                continue
            plazo = round((int('20' + year_match.group(1)) - datetime.now().year) + 0.5, 2)
            if plazo <= 0:
                continue
            tes_data.append({'name': label.replace('  ', ' '), 'tir': tir, 'plazo': plazo})

    tes_data.sort(key=lambda x: x['plazo'])
    return tes_data

def main():
    data = scrape()
    if not data:
        print("Sin datos - abortando")
        return
    result = {
        'fecha': datetime.now().strftime('%Y-%m-%d'),
        'fuente': 'BanRep SEN',
        'tes': data
    }
    with open('datos_curva.json', 'w') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)
    print(f"✓ {len(data)} TES guardados")
    for t in data:
        print(f"  {t['name']}: {t['tir']}% ({t['plazo']}a)")

if __name__ == '__main__':
    main()
