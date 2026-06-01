"""
scraper_eme.py — Encuesta Mensual de Expectativas BanRep (EME)
Busca el archivo más reciente en:
  https://www.banrep.gov.co/es/resultados-mensuales-expectativas-analistas-economicos
Guarda eme_data.json
"""
import json, re, io, datetime, requests
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

TODAY = datetime.date.today().isoformat()
NOW   = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

BASE_URL = "https://www.banrep.gov.co"
PAGE_URL = f"{BASE_URL}/es/resultados-mensuales-expectativas-analistas-economicos"

MESES_ES = ["ene", "feb", "mar", "abr", "may", "jun",
            "jul", "ago", "sep", "oct", "nov", "dic"]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
}


# ── DESCUBRIMIENTO DE URL ────────────────────────────────────────────────────

def find_url_from_page():
    """Parsea la página BanRep y retorna (url, filename) del Excel más reciente."""
    try:
        r = requests.get(PAGE_URL, headers=HEADERS, timeout=20)
        if r.status_code != 200:
            print(f"  Página retornó {r.status_code}")
            return None, None
        links = re.findall(r'/sites/default/files/(res_inf_\w+?\.xlsx)', r.text)
        if not links:
            print("  No se encontraron links en la página")
            return None, None
        filename = links[0]  # el más reciente aparece primero
        url = f"{BASE_URL}/sites/default/files/{filename}"
        print(f"  Encontrado en página: {filename}")
        return url, filename
    except Exception as e:
        print(f"  Error leyendo página: {e}")
        return None, None


def guess_url_by_month():
    """Prueba URLs mes a mes en reversa (hasta 6 meses atrás)."""
    d = datetime.date.today().replace(day=1)
    for _ in range(6):
        mes = MESES_ES[d.month - 1]
        filename = f"res_inf_{mes}{d.year}.xlsx"
        url = f"{BASE_URL}/sites/default/files/{filename}"
        try:
            r = requests.head(url, headers=HEADERS, timeout=10, allow_redirects=True)
            if r.status_code == 200:
                print(f"  URL válida: {filename}")
                return url, filename
        except Exception:
            pass
        # Retroceder un mes
        d = (d - datetime.timedelta(days=1)).replace(day=1)
    return None, None


# ── PARSEO DEL EXCEL ─────────────────────────────────────────────────────────

def parse_fecha(ws):
    """Extrae la frase de fecha de realización (fila 2)."""
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        txt = str(row[0] or "")
        return txt.replace("Fecha de realización:", "").replace("Fecha de realización:", "").strip()
    return ""


def parse_resumen(ws):
    """
    Lee el sheet RESUMEN y devuelve:
      inflacion_total, inflacion_sin_alimentos, trm
    Cada entrada: {horizonte, media, min, max, n}
    """
    inflacion_total = []
    inflacion_sin = []
    trm = []
    section = None

    for row in ws.iter_rows(values_only=True):
        label = str(row[0] or "").strip()
        val   = row[1]
        vmin  = row[2]
        vmax  = row[3]
        n     = row[4]

        lo = label.lower()

        # Detectar sección (orden importa: más específicos primero)
        if "cinco mejores" in lo:
            section = None; continue  # sub-sección "top 5" — ignorar
        if "política monetaria" in lo and val is None:
            break  # ya no hay más datos relevantes en RESUMEN
        if "sin alimentos ni regulados" in lo and val is None:
            section = "ipc_sar"; continue
        if "sin alimentos" in lo and val is None:
            section = "ipc_sin"; continue
        if "inflación total" in lo and val is None:
            section = "ipc_total"; continue
        if "de alimentos" in lo and val is None:
            section = "ipc_ali"; continue
        if "regulados" in lo and val is None:
            section = "ipc_reg"; continue
        if ("representativa" in lo or "trm" in lo) and val is None:
            section = "trm"; continue

        if not isinstance(val, float):
            continue

        entry = {
            "horizonte": label,
            "media": round(val * 100, 4),
            "min":   round((vmin or 0) * 100, 4),
            "max":   round((vmax or 0) * 100, 4),
            "n":     n,
        }

        if section == "ipc_total":
            inflacion_total.append(entry)
        elif section == "ipc_sin":
            inflacion_sin.append(entry)
        elif section == "trm" and val > 100:   # TRM en pesos, no porcentaje
            trm.append({
                "horizonte": label,
                "media": round(val, 0),
                "min":   round(vmin or 0, 0),
                "max":   round(vmax or 0, 0),
                "n":     n,
            })

    return inflacion_total, inflacion_sin, trm


def parse_tasa_interv(ws):
    """
    Lee TASA_INTERV y devuelve la senda esperada (todas las entidades).
    Estructura del sheet: Media/Mediana/Moda bajo 'De tendencia',
    Mínimo/Máximo bajo 'De dispersión' — ambas sub-secciones dentro de
    'TODAS LAS ENTIDADES PARTICIPANTES'.
    Retorna lista de {fecha, media, min, max}
    """
    dates = []
    in_all = False
    got_media = got_min = got_max = False
    path = []

    for row in ws.iter_rows(values_only=True):
        label = str(row[0] or "").strip()
        lo = label.lower()

        # Primera fila con fechas datetime → cabecera de columnas
        if not dates and any(isinstance(v, datetime.datetime) for v in row[1:]):
            dates = [
                v.strftime("%Y-%m-%d") if isinstance(v, datetime.datetime) else None
                for v in row[1:]
            ]
            continue

        if "todas las entidades" in lo:
            in_all = True; continue

        # Terminar al llegar a la siguiente sección (bancos, comisionistas…)
        if in_all and lo and not lo.startswith(" ") and "todas" not in lo and \
                any(k in lo for k in ("bancos", "comisionistas", "compa", "fondos")):
            break

        if not in_all:
            continue

        vals = list(row[1:])

        # Media (de tendencia)
        if "media" in lo and "moda" not in lo and "mediana" not in lo and not got_media:
            for i, d in enumerate(dates):
                if d and i < len(vals) and isinstance(vals[i], float):
                    path.append({"fecha": d, "media": round(vals[i] * 100, 4)})
            got_media = True

        # Mínimo (de dispersión — aparece después de tendencia)
        elif "nimo" in lo and got_media and not got_min:
            for i, p in enumerate(path):
                if i < len(vals) and isinstance(vals[i], float):
                    p["min"] = round(vals[i] * 100, 4)
            got_min = True

        # Máximo
        elif "ximo" in lo and got_media and not got_max:
            for i, p in enumerate(path):
                if i < len(vals) and isinstance(vals[i], float):
                    p["max"] = round(vals[i] * 100, 4)
            got_max = True

        if got_media and got_min and got_max:
            break

    return path


# ── MAIN ─────────────────────────────────────────────────────────────────────

def prev_month_url(filename):
    """Retorna (url, filename) del mes anterior al archivo dado."""
    m = re.search(r'res_inf_([a-z]+)(\d{4})\.xlsx', filename)
    if not m:
        return None, None
    mes_str, year = m.group(1), int(m.group(2))
    if mes_str not in MESES_ES:
        return None, None
    idx = MESES_ES.index(mes_str)
    prev_mes = MESES_ES[idx - 1] if idx > 0 else MESES_ES[11]
    prev_year = year if idx > 0 else year - 1
    fn = f"res_inf_{prev_mes}{prev_year}.xlsx"
    url = f"{BASE_URL}/sites/default/files/{fn}"
    try:
        r = requests.head(url, headers=HEADERS, timeout=10, allow_redirects=True)
        if r.status_code == 200:
            return url, fn
    except Exception:
        pass
    return None, None


def parse_and_save(url, filename, out_path, full=True):
    """Descarga, parsea y guarda un archivo EME. full=True incluye IPC/TRM."""
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    sheets = wb.sheetnames
    m = re.search(r'res_inf_([a-z]+)(\d{4})\.xlsx', filename)
    mes_label = (m.group(1) + " " + m.group(2)) if m else filename
    fecha_encuesta = parse_fecha(wb[sheets[0]])
    inflacion_total, inflacion_sin, trm = [], [], []
    if full and "RESUMEN" in sheets:
        inflacion_total, inflacion_sin, trm = parse_resumen(wb["RESUMEN"])
    tasa_path = []
    if "TASA_INTERV" in sheets:
        tasa_path = parse_tasa_interv(wb["TASA_INTERV"])
    result = {
        "updated":                 NOW,
        "source":                  f"BanRep EME · {mes_label}",
        "fecha_encuesta":          fecha_encuesta,
        "inflacion_total":         inflacion_total,
        "inflacion_sin_alimentos": inflacion_sin,
        "trm":                     trm,
        "tasa_intervencion":       tasa_path,
    }
    Path(out_path).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK {out_path} guardado ({len(tasa_path)} reuniones en senda)")
    return result


def main():
    print(f"=== EME BanRep · {TODAY} ===\n")

    # 1. Localizar mes actual
    print("Buscando último archivo EME en página BanRep...")
    url, filename = find_url_from_page()
    if not url:
        print("Intentando por mes en reversa...")
        url, filename = guess_url_by_month()

    if not url:
        print("✗ No se encontró archivo EME.")
        if Path("eme_data.json").exists():
            print("  Manteniendo eme_data.json existente.")
        return

    # 2. Descargar y guardar mes actual (completo)
    print(f"\nDescargando {filename}...")
    try:
        result = parse_and_save(url, filename, "eme_data.json", full=True)
        if result["inflacion_total"]:
            print(f"  IPC total ({len(result['inflacion_total'])} horizontes):")
            for e in result["inflacion_total"][:3]:
                print(f"    {e['horizonte']}: {e['media']:.2f}%")
        if result["trm"]:
            print(f"  TRM: {result['trm'][0]['media']:,.0f}")
    except Exception as e:
        print(f"✗ Error mes actual: {e}")
        return

    # 3. Descargar mes anterior (solo senda de tasa para comparación)
    print(f"\nBuscando mes anterior a {filename}...")
    url_prev, filename_prev = prev_month_url(filename)
    if url_prev:
        print(f"Descargando {filename_prev}...")
        try:
            parse_and_save(url_prev, filename_prev, "eme_data_prev.json", full=False)
        except Exception as e:
            print(f"  Error mes anterior: {e}")
    else:
        print("  No se encontró mes anterior.")


if __name__ == "__main__":
    main()
